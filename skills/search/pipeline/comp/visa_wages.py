"""US visa-wage compensation backend (DOL Labor Condition Application disclosure).

Two concerns live here:

* ``lookup`` — exact-normalised employer match against the local visa-wage SQLite
  dataset (built by the refresh CLI). A missing/stale dataset degrades to
  ``None`` (the pipeline then skips this source with a console note).
* ``--refresh`` CLI (``python -m pipeline.comp.visa_wages --refresh``) — a
  separate, manual acquisition step that discovers the newest DOL visa-wage
  quarterly ``.xlsx`` from the OFLC performance-page index, downloads it (DOL
  sits behind Akamai; plain ``requests`` gets a 403, so we use ``curl_cffi``),
  streams it into ``.agents/search/visa-wages/wage-index.sqlite``, and writes
  ``meta.json``.

Neither part runs inside a normal search run. ``curl_cffi``/``openpyxl`` are
imported lazily inside the refresh path; importing this module makes no network
calls.
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from pipeline.config import state_dir
from pipeline.model import CompRecord

_PERFORMANCE_PAGE = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
_ARCHIVES_PREFIX = "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/"
_TABLE = "wages"
_DB_NAME = "wage-index.sqlite"
_META_NAME = "meta.json"

# A copied visa-wage dataset is "stale" after about one quarter.
_STALE_DAYS = 120

# LCA disclosure files follow ``LCA_<Name>_Data_FY<yyyy>_Q<q>.xlsx``. The
# current-quarter file on the DOL page carries a typo in ``<Name>``
# (``LCA_Dislclosure_...``), which the ``[A-Za-z]+`` segment tolerates.
_FILE_RE = re.compile(r"LCA_[A-Za-z]+_Data_FY(\d{4})_Q([1-4])\.xlsx", re.IGNORECASE)

# Corporate suffixes stripped for exact-normalised employer matching.
_SUFFIXES = ("incorporated", "corporation", "inc", "llc", "corp")


def _visa_wages_dir() -> Path:
    """The visa-wage data dir: ``$SEARCH_STATE_DIR/visa-wages``."""
    return state_dir() / "visa-wages"


# ---- normalisation (shared by lookup and refresh) --------------------------


def normalize_employer(name: str) -> str:
    """Normalise an employer name for exact matching.

    Lowercase, strip every non-alphanumeric, then drop a trailing
    ``Inc|LLC|Corp`` (or its spelled-out form). Used on the query side by
    ``lookup`` and on the store side by ``refresh``.
    """
    norm = re.sub(r"[^a-z0-9]", "", (name or "").lower())
    for suffix in _SUFFIXES:
        if norm.endswith(suffix):
            norm = norm[: -len(suffix)]
            break
    return norm


# ---- lookup ----------------------------------------------------------------


def _read_fetched_at(meta_path: Path) -> datetime | None:
    try:
        data = json.loads(meta_path.read_text(encoding="utf-8"))
        raw = data.get("fetched_at")
        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except (OSError, ValueError, TypeError, KeyError, json.JSONDecodeError):
        return None


def _is_stale(fetched_at: datetime, now: datetime | None = None) -> bool:
    now = now if now is not None else datetime.now(timezone.utc)
    return (now - fetched_at).total_seconds() > _STALE_DAYS * 86400


def lookup(company: str, sqlite_path: Path) -> CompRecord | None:
    """Exact-normalised employer match against the local visa-wage dataset.

    Returns a ladder-free :class:`CompRecord` with ``floor_value`` = the lowest
    ``annual_base`` among the employer's rows, ``provenance="visa-wages"`` and
    ``employer`` = the matched name. Missing sqlite/meta (or a stale copy)
    degrades to ``None`` — never raises.
    """
    sqlite_path = Path(sqlite_path)
    meta_path = sqlite_path.parent / _META_NAME
    if not sqlite_path.is_file() or not meta_path.is_file():
        return None

    fetched_at = _read_fetched_at(meta_path)
    if fetched_at is None:
        return None
    if _is_stale(fetched_at):
        print(
            f"[visa-wages] dataset at {sqlite_path} is stale "
            f"(fetched {fetched_at.date().isoformat()}); skipping visa-wage lookup",
            file=sys.stderr,
        )
        return None

    norm = normalize_employer(company)
    if not norm:
        return None

    try:
        conn = sqlite3.connect(str(sqlite_path))
        try:
            row = conn.execute(
                f"SELECT MIN(annual_base) FROM {_TABLE} "
                "WHERE employer = ? AND annual_base IS NOT NULL",
                (norm,),
            ).fetchone()
        finally:
            conn.close()
    except sqlite3.Error:
        return None

    if row is None or row[0] is None:
        return None
    return CompRecord(
        provenance="visa-wages",
        floor_value=float(row[0]),
        currency="USD",
        ladder=None,
        range=None,
        employer=company,
    )


# ---- refresh CLI (separate acquisition step) -------------------------------


def _parse_quarter(filename: str) -> tuple[int, int] | None:
    m = re.search(r"FY(\d{4})_Q([1-4])", filename, re.IGNORECASE)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _newest_file(html: str) -> tuple[str | None, str | None]:
    """Return ``(filename, quarter_label)`` for the newest LCA disclosure file
    referenced by the performance-page index. The discovered filename is used
    verbatim (typo included) when constructing the archives URL."""
    candidates: dict[tuple[int, int], str] = {}
    for match in _FILE_RE.finditer(html or ""):
        name = match.group(0)
        quarter = _parse_quarter(name)
        if quarter:
            candidates.setdefault(quarter, name)
    if not candidates:
        return None, None
    fy, qtr = max(candidates)
    return candidates[(fy, qtr)], f"FY{fy}_Q{qtr}"


def _is_xlsx(path: Path) -> bool:
    try:
        with open(path, "rb") as fh:
            return fh.read(4) == b"PK\x03\x04"
    except OSError:
        return False


def _download(filename: str, visa_wages_dir: Path) -> Path | None:
    """Download ``filename`` into ``visa_wages_dir`` via ``curl_cffi``.

    The current-quarter URL on the index has a typo and a different path
    prefix, so we construct the canonical archives URL from the discovered
    filename and fall back to a couple of variants (including the corrected
    spelling) if the primary candidate 404s.
    """
    from curl_cffi import requests

    corrected = filename.replace("Dislclosure", "Disclosure")
    candidates = [
        _ARCHIVES_PREFIX + filename,
        f"https://www.dol.gov/sites/dolgov/files/ETA/oflc/{filename}",
        _ARCHIVES_PREFIX + corrected,
        f"https://www.dol.gov/sites/dolgov/files/ETA/oflc/{corrected}",
    ]
    dest = visa_wages_dir / filename
    for url in candidates:
        try:
            with requests.get(url, stream=True, impersonate="chrome", timeout=180) as resp:
                if resp.status_code != 200:
                    continue
                with open(dest, "wb") as fh:
                    for chunk in resp.iter_content(chunk_size=1 << 16):
                        if chunk:
                            fh.write(chunk)
            if _is_xlsx(dest):
                return dest
            dest.unlink(missing_ok=True)
        except Exception:  # noqa: BLE001 — try the next candidate
            dest.unlink(missing_ok=True)
            continue
    return None


_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "employer": ("LCACASEEMPLOYERNAME", "EMPLOYERNAME", "EMPLOYERLEGALBUSINESSNAME", "EMPLOYER"),
    "state": (
        "LCACASEWORKLOC1STATE",
        "WORKLOC1STATE",
        "WORKLOCSTATE",
        "WORKSITESTATE",
        "LCACASEEMPLOYERSTATE",
        "EMPLOYERSTATE",
    ),
    "job_title": ("LCACASEJOBTITLE", "JOBTITLE"),
    "soc_code": ("LCACASESOCCODE", "SOCCODE"),
    "soc_title": ("LCACASESOCNAME", "SOCNAME", "SOCTITLE"),
    "wage_level": ("PWWAGELEVEL", "WAGELEVEL"),
    "wage_from": (
        "LCACASEWAGERATEFROM",
        "WAGERATEFROM",
        "WAGERATEOFPAYFROM",
    ),
    "wage_unit": (
        "LCACASEWAGERATEUNIT",
        "WAGERATEUNIT",
        "WAGEUNITOFPAY",
    ),
}

_WAGE_UNIT_FACTORS: dict[str, float] = {
    "year": 1.0, "yr": 1.0, "annual": 1.0, "annually": 1.0,
    "hour": 2080.0, "hr": 2080.0, "hourly": 2080.0,
    "week": 52.0, "wk": 52.0, "weekly": 52.0,
    "bi-weekly": 26.0, "biweekly": 26.0, "bi weekly": 26.0,
    "month": 12.0, "mo": 12.0, "monthly": 12.0,
}


def _column_indexes(header: tuple) -> dict[str, int]:
    normalised: dict[str, int] = {}
    for i, name in enumerate(header or ()):
        if isinstance(name, str):
            normalised.setdefault(re.sub(r"[^A-Z0-9]", "", name.upper()), i)
    indexes: dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                indexes[field] = normalised[alias]
                break
    return indexes


def _cell(row: tuple, idx: int | None) -> str | None:
    if idx is None or row is None or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    text = str(value).strip()
    return text or None


def _annualise(wage_from: str | None, wage_unit: str | None) -> float | None:
    """Annualise a wage figure using its unit; unknown units yield ``None``."""
    if wage_from is None:
        return None
    try:
        value = float(str(wage_from).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None
    unit = (wage_unit or "year").strip().lower()
    factor = _WAGE_UNIT_FACTORS.get(unit)
    if factor is None:
        return None
    return value * factor


def _stream_into_sqlite(xlsx_path: Path, db_path: Path) -> int:
    """Stream the xlsx into sqlite, skipping padding rows (empty employer or
    job title). Returns the number of rows written."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        # Some files ship a cover sheet first; pick the first sheet whose
        # header carries the columns we need.
        target_ws = None
        indexes: dict[str, int] = {}
        for ws in wb.worksheets:
            header = next(ws.iter_rows(values_only=True), None)
            if header is None:
                continue
            cand = _column_indexes(header)
            if {"employer", "job_title", "wage_from"} <= cand.keys():
                target_ws = ws
                indexes = cand
                break
        if target_ws is None:
            return 0

        conn = sqlite3.connect(str(db_path))
        try:
            conn.execute(
                f"CREATE TABLE IF NOT EXISTS {_TABLE} ("
                " employer TEXT NOT NULL,"
                " state TEXT,"
                " job_title TEXT,"
                " soc_code TEXT,"
                " soc_title TEXT,"
                " pw_wage_level TEXT,"
                " annual_base REAL"
                ")"
            )
            conn.execute(
                f"CREATE INDEX IF NOT EXISTS idx_{_TABLE}_employer ON {_TABLE}(employer)"
            )
            conn.execute(f"DELETE FROM {_TABLE}")

            count = 0
            rows = target_ws.iter_rows(values_only=True)
            next(rows, None)  # header
            for row in rows:
                if row is None:
                    continue
                employer_raw = _cell(row, indexes.get("employer"))
                job_title = _cell(row, indexes.get("job_title"))
                if not employer_raw or not job_title:
                    continue  # padding row
                employer = normalize_employer(employer_raw)
                if not employer:
                    continue
                conn.execute(
                    f"INSERT INTO {_TABLE} "
                    "(employer, state, job_title, soc_code, soc_title, pw_wage_level, annual_base) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (
                        employer,
                        _cell(row, indexes.get("state")),
                        job_title,
                        _cell(row, indexes.get("soc_code")),
                        _cell(row, indexes.get("soc_title")),
                        _cell(row, indexes.get("wage_level")),
                        _annualise(
                            _cell(row, indexes.get("wage_from")),
                            _cell(row, indexes.get("wage_unit")),
                        ),
                    ),
                )
                count += 1
            conn.commit()
            return count
        finally:
            conn.close()
    finally:
        wb.close()


def _write_meta(meta_path: Path, quarter: str) -> None:
    meta_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "quarter": quarter,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
    }
    meta_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _refresh(visa_wages_dir: Path) -> int:
    try:
        import curl_cffi  # noqa: F401
    except ImportError:
        print("visa-wages --refresh requires curl-cffi (pip install curl-cffi)", file=sys.stderr)
        return 1

    visa_wages_dir.mkdir(parents=True, exist_ok=True)
    from curl_cffi import requests

    print(f"[visa-wages] discovering newest visa-wage file from {_PERFORMANCE_PAGE}", file=sys.stderr)
    try:
        resp = requests.get(_PERFORMANCE_PAGE, impersonate="chrome", timeout=60)
        resp.raise_for_status()
        html = resp.text
    except Exception as exc:  # noqa: BLE001
        print(f"[visa-wages] could not fetch the performance page: {exc}", file=sys.stderr)
        return 1

    filename, quarter = _newest_file(html)
    if filename is None:
        print("[visa-wages] could not discover a disclosure file on the performance page", file=sys.stderr)
        return 1

    print(f"[visa-wages] downloading {filename} ({quarter})", file=sys.stderr)
    xlsx_path = _download(filename, visa_wages_dir)
    if xlsx_path is None:
        print(f"[visa-wages] download failed for {filename}", file=sys.stderr)
        return 1

    db_path = visa_wages_dir / _DB_NAME
    rows = _stream_into_sqlite(xlsx_path, db_path)
    _write_meta(visa_wages_dir / _META_NAME, quarter)
    xlsx_path.unlink(missing_ok=True)
    print(f"[visa-wages] refreshed {quarter}: {rows} rows -> {db_path}", file=sys.stderr)
    return 0


def main(argv: list[str] | None = None) -> int:
    """CLI entry point: ``python -m pipeline.comp.visa_wages --refresh``."""
    import argparse

    parser = argparse.ArgumentParser(prog="python -m pipeline.comp.visa_wages")
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="download the newest visa-wage quarterly file and rebuild the local sqlite",
    )
    args = parser.parse_args(argv)
    if args.refresh:
        return _refresh(_visa_wages_dir())
    parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
