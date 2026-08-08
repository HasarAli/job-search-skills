#!/usr/bin/env python3
"""Build the local DOL LCA salary index (SQLite) from OFLC quarterly disclosure files.

Discovers the newest N quarter XLSX files from the OFLC performance page, downloads
any not already cached, streams each row-by-row (openpyxl read_only, so a 140MB file
never loads whole into RAM), keeps CERTIFIED filings, annualizes the wage, and writes
one indexed SQLite table. Idempotent: re-run per quarter; cached XLSX are reused.

  --quarters N   number of most-recent quarters to index (default 4)
  --rebuild      drop and rebuild the table (default: rebuild is always full)
  --keep-xlsx    keep downloaded XLSX in the cache (default: keep, for re-runs)

Deterministic: no LLM, no network at query time. Requires curl_cffi (DOL is behind
Akamai bot protection that blocks plain urllib/requests by TLS fingerprint).
"""
import argparse
import re
import sqlite3
import sys

import openpyxl
from curl_cffi import requests

from salary_common import (
    CACHE_DIR, DB_PATH, DOL_HOST, KEEP_COLS, PERF_PAGE,
    annualize, norm_employer, to_float,
)

_QUARTER_RE = re.compile(r"LCA_Disclosure_Data_FY(\d{4})_Q(\d)\.xlsx", re.I)


def newest_quarter_urls(n: int) -> list[tuple[str, str]]:
    """Return [(quarter_label, full_url)] for the n most recent quarters, newest last."""
    r = requests.get(PERF_PAGE, impersonate="chrome", timeout=60)
    r.raise_for_status()
    found = {}
    for href in re.findall(r'href="([^"]+LCA_Disclosure_Data[^"]+\.xlsx)"', r.text, re.I):
        m = _QUARTER_RE.search(href)
        if m:
            key = (int(m.group(1)), int(m.group(2)))  # (FY, Q) for correct numeric sort
            url = href if href.startswith("http") else DOL_HOST + href
            found[key] = (f"FY{m.group(1)}_Q{m.group(2)}", url)
    ordered = [found[k] for k in sorted(found)]
    return ordered[-n:]


def download(url: str, dest) -> None:
    if dest.exists() and dest.stat().st_size > 0:
        print(f"  cached {dest.name}", file=sys.stderr)
        return
    print(f"  downloading {dest.name} ...", file=sys.stderr)
    r = requests.get(url, impersonate="chrome", timeout=600)
    r.raise_for_status()
    dest.write_bytes(r.content)
    print(f"  saved {dest.name} ({len(r.content)//1_000_000} MB)", file=sys.stderr)


def stream_rows(xlsx_path, quarter, cur):
    """Stream one XLSX into the DB. Returns (scanned, inserted)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {name: i for i, name in enumerate(header)}
    missing = [c for c in KEEP_COLS if c not in idx]
    if missing:
        raise KeyError(f"{xlsx_path.name} missing columns {missing} — schema changed")

    def g(row, col):
        return row[idx[col]]

    def txt(v):  # cells can arrive as int/float (numeric-looking strings); coerce before .strip()
        return "" if v is None else str(v).strip()

    scanned = inserted = 0
    batch = []
    for row in it:
        scanned += 1
        if str(g(row, "CASE_STATUS") or "").strip().lower() != "certified":
            continue
        lo = annualize(to_float(g(row, "WAGE_RATE_OF_PAY_FROM")), g(row, "WAGE_UNIT_OF_PAY"))
        if lo is None:
            continue  # unusable wage -> drop; base salary is the whole point
        hi = annualize(to_float(g(row, "WAGE_RATE_OF_PAY_TO")), g(row, "WAGE_UNIT_OF_PAY")) or lo
        emp = txt(g(row, "EMPLOYER_NAME"))
        batch.append((
            norm_employer(emp), emp, txt(g(row, "JOB_TITLE")),
            txt(g(row, "SOC_CODE")), txt(g(row, "SOC_TITLE")),
            lo, hi, annualize(to_float(g(row, "PREVAILING_WAGE")), "year"),
            txt(g(row, "PW_WAGE_LEVEL")),
            txt(g(row, "WORKSITE_CITY")), txt(g(row, "WORKSITE_STATE")),
            txt(g(row, "FULL_TIME_POSITION")), txt(g(row, "BEGIN_DATE"))[:10],
            quarter,
        ))
        inserted += 1
        if len(batch) >= 5000:
            cur.executemany(_INSERT, batch)
            batch.clear()
        if scanned % 100_000 == 0:
            print(f"    {quarter}: scanned {scanned:,} kept {inserted:,}", file=sys.stderr)
    if batch:
        cur.executemany(_INSERT, batch)
    wb.close()
    return scanned, inserted


_INSERT = """INSERT INTO filings(
    employer_norm, employer_name, job_title, soc_code, soc_title,
    wage_from_annual, wage_to_annual, prevailing_annual, pw_level,
    city, state, full_time, begin_date, quarter
) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--quarters", type=int, default=4)
    args = ap.parse_args()

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Resolving newest {args.quarters} quarters ...", file=sys.stderr)
    quarters = newest_quarter_urls(args.quarters)
    print("Indexing:", ", ".join(q for q, _ in quarters), file=sys.stderr)

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.executescript("""
        DROP TABLE IF EXISTS filings;
        CREATE TABLE filings(
            employer_norm TEXT, employer_name TEXT, job_title TEXT,
            soc_code TEXT, soc_title TEXT,
            wage_from_annual REAL, wage_to_annual REAL, prevailing_annual REAL,
            pw_level TEXT, city TEXT, state TEXT, full_time TEXT,
            begin_date TEXT, quarter TEXT
        );
    """)

    total_scanned = total_kept = 0
    for label, url in quarters:
        dest = CACHE_DIR / f"LCA_Disclosure_Data_{label}.xlsx"
        download(url, dest)
        s, k = stream_rows(dest, label, cur)
        con.commit()
        total_scanned += s
        total_kept += k
        print(f"  {label}: {k:,} certified rows indexed", file=sys.stderr)

    print("Building indexes ...", file=sys.stderr)
    cur.executescript("""
        CREATE INDEX ix_emp ON filings(employer_norm);
        CREATE INDEX ix_soc ON filings(soc_code);
    """)
    con.commit()
    con.close()
    print(f"Done. {total_kept:,} rows from {total_scanned:,} scanned -> {DB_PATH}", file=sys.stderr)


if __name__ == "__main__":
    main()
